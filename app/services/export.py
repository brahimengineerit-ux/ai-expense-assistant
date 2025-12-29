"""
Export Service
==============
Generate Excel and CSV exports from expense data.
"""

import csv
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_to_csv(expenses: list[dict]) -> str:
    """
    Export expenses to CSV format.
    
    Args:
        expenses: List of expense dictionaries
    
    Returns:
        CSV string
    """
    if not expenses:
        return ""
    
    # Get all unique keys
    all_keys = set()
    for expense in expenses:
        all_keys.update(expense.keys())
    
    fieldnames = sorted(all_keys)
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(expenses)
    
    return output.getvalue()


def export_to_excel(
    expenses: list[dict],
    include_summary: bool = True,
    title: str = "Expense Report"
) -> bytes:
    """
    Export expenses to Excel with formatting.
    
    Args:
        expenses: List of expense dictionaries
        include_summary: Whether to add summary sheet
        title: Report title
    
    Returns:
        Excel file bytes
    """
    wb = Workbook()
    
    # ========== EXPENSES SHEET ==========
    ws = wb.active
    ws.title = "Expenses"
    
    if not expenses:
        ws["A1"] = "No expenses to export"
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    # Get headers
    headers = ["amount", "currency", "category", "description", "date", "payment_method"]
    all_keys = set()
    for expense in expenses:
        all_keys.update(expense.keys())
    
    # Add any extra keys not in default headers
    for key in sorted(all_keys):
        if key not in headers:
            headers.append(key)
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Title
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color="666666")
    
    # Headers (row 4)
    header_row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data
    for row_idx, expense in enumerate(expenses, header_row + 1):
        for col_idx, header in enumerate(headers, 1):
            value = expense.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # Format amounts
            if header == "amount" and isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        
        for row in range(header_row, len(expenses) + header_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        ws.column_dimensions[column_letter].width = min(max_length + 2, 30)
    
    # ========== SUMMARY SHEET ==========
    if include_summary:
        ws_summary = wb.create_sheet(title="Summary")
        
        # Calculate summary data
        total = sum(e.get("amount", 0) or 0 for e in expenses)
        categories = {}
        payment_methods = {}
        
        for expense in expenses:
            cat = expense.get("category") or "Other"
            pm = expense.get("payment_method") or "Unknown"
            amount = expense.get("amount") or 0
            
            categories[cat] = categories.get(cat, 0) + amount
            payment_methods[pm] = payment_methods.get(pm, 0) + amount
        
        # Title
        ws_summary["A1"] = "Expense Summary"
        ws_summary["A1"].font = Font(bold=True, size=16)
        
        # Overview
        ws_summary["A3"] = "Overview"
        ws_summary["A3"].font = Font(bold=True, size=12)
        ws_summary["A4"] = "Total Expenses"
        ws_summary["B4"] = len(expenses)
        ws_summary["A5"] = "Total Amount"
        ws_summary["B5"] = total
        ws_summary["B5"].number_format = "#,##0.00"
        
        # By Category
        ws_summary["A7"] = "By Category"
        ws_summary["A7"].font = Font(bold=True, size=12)
        row = 8
        for cat, amount in sorted(categories.items(), key=lambda x: -x[1]):
            ws_summary.cell(row=row, column=1, value=cat)
            ws_summary.cell(row=row, column=2, value=amount).number_format = "#,##0.00"
            pct = (amount / total * 100) if total > 0 else 0
            ws_summary.cell(row=row, column=3, value=f"{pct:.1f}%")
            row += 1
        
        # By Payment Method
        row += 1
        ws_summary.cell(row=row, column=1, value="By Payment Method").font = Font(bold=True, size=12)
        row += 1
        for pm, amount in sorted(payment_methods.items(), key=lambda x: -x[1]):
            ws_summary.cell(row=row, column=1, value=pm)
            ws_summary.cell(row=row, column=2, value=amount).number_format = "#,##0.00"
            row += 1
        
        # Adjust widths
        ws_summary.column_dimensions["A"].width = 20
        ws_summary.column_dimensions["B"].width = 15
        ws_summary.column_dimensions["C"].width = 10
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
